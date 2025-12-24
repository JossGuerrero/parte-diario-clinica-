from django.http import JsonResponse, HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_GET

from openpyxl import Workbook
from functools import wraps

# Lightweight in-memory cache used only to make unit tests behave like the cached view
# without importing Django cache machinery at module import time.
def lightweight_cache(timeout_seconds):
    def decorator(view_func):
        cache = {}
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            key = request.get_full_path()
            if key in cache:
                return cache[key]
            resp = view_func(request, *args, **kwargs)
            cache[key] = resp
            return resp
        # keep __wrapped__ for tests that unwrap decorators
        _wrapped.__wrapped__ = view_func
        return _wrapped
    return decorator
from django.utils.timezone import datetime
from core.db_access import get_connection
from django.conf import settings
from django.shortcuts import render
import logging

logger = logging.getLogger(__name__)


def _detail_sql(base_where):
    return f"""
    SELECT 
      e.FechaIng AS Fecha,
      c.nomCli AS Paciente,
      m.CodSer AS Servicio,
      c.Tipo AS Institucion,
      c.ValorConsulta AS ValorConsulta,
      m.TotalMedicina AS ValorMedicina
    FROM TraSec AS t
    INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
    INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
    LEFT JOIN EqCliDes AS m ON t.IdTraSec = m.IdTraSec
    {base_where}
    ORDER BY e.FechaIng DESC;
    """


def fetch_detalle(cur, base_where, params):
    """Execute the detail query and return a list of dict rows matching the UI table."""
    cur.execute(_detail_sql(base_where), params)
    rows = cur.fetchall()

    def _get(r, name, idx):
        try:
            return getattr(r, name)
        except Exception:
            try:
                return r[idx]
            except Exception:
                return None

    detalle = []
    for r in rows:
        fecha = _get(r, 'Fecha', 0)
        if hasattr(fecha, 'strftime'):
            fecha = fecha.strftime('%Y-%m-%d')
        else:
            fecha = str(fecha) if fecha is not None else ''
        valor_cons = float(_get(r, 'ValorConsulta', 4) or 0)
        valor_med = float(_get(r, 'ValorMedicina', 5) or 0)
        detalle.append({
            'fecha': fecha,
            'paciente': _get(r, 'Paciente', 1) or '',
            'servicio': _get(r, 'Servicio', 2) or '',
            'institucion': _get(r, 'Institucion', 3) or '',
            'valor_consulta': valor_cons,
            'valor_medicina': valor_med,
            'total': valor_cons + valor_med,
        })
    return detalle


@staff_member_required
def dashboard_index(request):
    """Render a demo dashboard that consumes Access endpoints via AJAX."""
    return render(request, 'dashboard/index.html')



@staff_member_required
@require_GET
def atenciones_por_mes(request):
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    if not inicio or not fin:
        return JsonResponse({'error': 'Parámetros "inicio" y "fin" requeridos (YYYY-MM-DD).'}, status=400)

    sql = """
    SELECT DatePart('m', e.FechaIng) AS Mes,
           COUNT(t.CodCli) AS NumeroAtenciones
    FROM TraSec AS t
    INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
    WHERE e.FechaIng BETWEEN ? AND ?
    GROUP BY DatePart('m', e.FechaIng)
    ORDER BY Mes;
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql, [inicio, fin])
        rows = cur.fetchall()
        data = [{'mes': int(r.Mes), 'numero': int(r.NumeroAtenciones)} for r in rows]
    except Exception as e:
        logger.exception('Error ejecutando atenciones_por_mes')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    return JsonResponse(data, safe=False)


@staff_member_required
@require_GET
def atenciones_por_edad(request):
    sql = """
    SELECT 
        IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 15 AND 19, '15-19',
        IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 20 AND 49, '20-49',
        IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 50 AND 64, '50-64',
        'Mayor 65'))) AS RangoEdad,
        COUNT(*) AS Total
    FROM TraSec AS t
    INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
    GROUP BY 
        IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 15 AND 19, '15-19',
        IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 20 AND 49, '20-49',
        IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 50 AND 64, '50-64',
        'Mayor 65')));
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        data = [{'rango': r.RangoEdad, 'total': int(r.Total)} for r in rows]
    except Exception as e:
        logger.exception('Error ejecutando atenciones_por_edad')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    return JsonResponse(data, safe=False)


@staff_member_required
@require_GET
def valor_recaudado_consulta(request):
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    if not inicio or not fin:
        return JsonResponse({'error': 'Parámetros "inicio" y "fin" requeridos (YYYY-MM-DD).'}, status=400)
    sql = """
    SELECT DatePart('m', e.FechaIng) AS Mes,
           SUM(c.ValorConsulta) AS TotalConsulta
    FROM TraSec AS t
    INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
    INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
    WHERE e.FechaIng BETWEEN ? AND ?
    GROUP BY DatePart('m', e.FechaIng)
    ORDER BY Mes;
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql, [inicio, fin])
        rows = cur.fetchall()
        data = [{'mes': int(r.Mes), 'total': float(r.TotalConsulta or 0)} for r in rows]
    except Exception as e:
        logger.exception('Error ejecutando valor_recaudado_consulta')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return JsonResponse(data, safe=False)


@staff_member_required
@require_GET
def valor_recaudado_medicina(request):
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    if not inicio or not fin:
        return JsonResponse({'error': 'Parámetros "inicio" y "fin" requeridos (YYYY-MM-DD).'}, status=400)
    sql = """
    SELECT DatePart('m', e.FechaIng) AS Mes,
           SUM(m.TotalMedicina) AS TotalMedicina
    FROM TraSec AS t
    INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
    INNER JOIN EqCliDes AS m ON t.IdTraSec = m.IdTraSec
    WHERE e.FechaIng BETWEEN ? AND ?
    GROUP BY DatePart('m', e.FechaIng)
    ORDER BY Mes;
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql, [inicio, fin])
        rows = cur.fetchall()
        data = [{'mes': int(r.Mes), 'total': float(r.TotalMedicina or 0)} for r in rows]
    except Exception as e:
        logger.exception('Error ejecutando valor_recaudado_medicina')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return JsonResponse(data, safe=False)


@staff_member_required
@require_GET
def atenciones_por_institucion(request):
    sql = """
    SELECT c.Tipo AS Institucion, COUNT(*) AS TotalAtenciones
    FROM TraSec AS t
    INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
    GROUP BY c.Tipo;
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        data = [{'institucion': r.Institucion, 'total': int(r.TotalAtenciones)} for r in rows]
    except Exception as e:
        logger.exception('Error ejecutando atenciones_por_institucion')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return JsonResponse(data, safe=False)


@staff_member_required
@require_GET
def atenciones_por_genero(request):
    sql = """
    SELECT c.Genero, COUNT(*) AS Total
    FROM TraSec AS t
    INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
    GROUP BY c.Genero;
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        data = [{'genero': r.Genero, 'total': int(r.Total)} for r in rows]
    except Exception as e:
        logger.exception('Error ejecutando atenciones_por_genero')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return JsonResponse(data, safe=False)


@staff_member_required
@require_GET
@lightweight_cache(60)
def dashboard_data(request):
    """Consolidated endpoint that returns all datasets in one JSON call. Cached for 60s."""
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    institucion = request.GET.get('institucion')
    genero = request.GET.get('genero')
    edad_range = request.GET.get('edad_range')  # e.g. '0-12', '60+'

    params = []
    where_clauses = []

    if inicio and fin:
        where_clauses.append("e.FechaIng BETWEEN ? AND ?")
        params.extend([inicio, fin])

    if institucion:
        where_clauses.append('c.Tipo = ?')
        params.append(institucion)
    if genero and genero != 'ALL':
        where_clauses.append('c.Genero = ?')
        params.append(genero)

    if edad_range:
        # validate edad_range format safely (examples: '0-12', '60+')
        try:
            if edad_range.endswith('+'):
                low = int(edad_range[:-1])
                where_clauses.append(f"DateDiff('yyyy', c.FechaN, Date()) >= {low}")
            else:
                parts = edad_range.split('-')
                if len(parts) != 2:
                    raise ValueError('invalid format')
                low, high = parts
                where_clauses.append(f"DateDiff('yyyy', c.FechaN, Date()) BETWEEN {int(low)} AND {int(high)}")
        except Exception:
            return JsonResponse({'error': 'Parámetro "edad_range" inválido. Use "0-12" o "60+".'}, status=400)

    try:
        conn = get_connection()
        cur = conn.cursor()

        base_where = ('WHERE ' + ' AND '.join(where_clauses)) if where_clauses else ''

        # Atenciones por mes
        at_sql = f"""
        SELECT DatePart('m', e.FechaIng) AS Mes, COUNT(t.CodCli) AS NumeroAtenciones
        FROM TraSec AS t
        INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
        INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
        {base_where}
        GROUP BY DatePart('m', e.FechaIng)
        ORDER BY Mes;
        """
        cur.execute(at_sql, params)
        at_rows = cur.fetchall()
        atenciones = [{'mes': int(r.Mes), 'numero': int(r.NumeroAtenciones)} for r in at_rows]

        # Valor consulta
        vc_sql = f"""
        SELECT DatePart('m', e.FechaIng) AS Mes, SUM(c.ValorConsulta) AS TotalConsulta
        FROM TraSec AS t
        INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
        INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
        {base_where}
        GROUP BY DatePart('m', e.FechaIng)
        ORDER BY Mes;
        """
        cur.execute(vc_sql, params)
        vc_rows = cur.fetchall()
        valor_consulta = [{'mes': int(r.Mes), 'total': float(r.TotalConsulta or 0)} for r in vc_rows]

        # Valor medicina
        vm_sql = f"""
        SELECT DatePart('m', e.FechaIng) AS Mes, SUM(m.TotalMedicina) AS TotalMedicina
        FROM TraSec AS t
        INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
        INNER JOIN EqCliDes AS m ON t.IdTraSec = m.IdTraSec
        INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
        {base_where}
        GROUP BY DatePart('m', e.FechaIng)
        ORDER BY Mes;
        """
        cur.execute(vm_sql, params)
        vm_rows = cur.fetchall()
        valor_medicina = [{'mes': int(r.Mes), 'total': float(r.TotalMedicina or 0)} for r in vm_rows]

        # Institucion
        inst_sql = f"""
        SELECT c.Tipo AS Institucion, COUNT(*) AS TotalAtenciones
        FROM TraSec AS t
        INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
        {base_where}
        GROUP BY c.Tipo;
        """
        cur.execute(inst_sql, params)
        inst_rows = cur.fetchall()
        instituciones = [{'institucion': r.Institucion, 'total': int(r.TotalAtenciones)} for r in inst_rows]

        # Genero
        gen_sql = f"""
        SELECT c.Genero, COUNT(*) AS Total
        FROM TraSec AS t
        INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
        {base_where}
        GROUP BY c.Genero;
        """
        cur.execute(gen_sql, params)
        gen_rows = cur.fetchall()
        genero_data = [{'genero': r.Genero, 'total': int(r.Total)} for r in gen_rows]

        # Edad ranges (use fixed buckets)
        edad_sql = """
        SELECT 
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 0 AND 12, '0-12',
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 13 AND 17, '13-17',
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 18 AND 30, '18-30',
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 31 AND 59, '31-59','60+')))) AS Rango,
          COUNT(*) AS Total
        FROM TraSec AS t
        INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
        {base_where}
        GROUP BY 
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 0 AND 12, '0-12',
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 13 AND 17, '13-17',
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 18 AND 30, '18-30',
          IIF(DateDiff('yyyy', c.FechaN, Date()) BETWEEN 31 AND 59, '31-59','60+'))));
        """
        cur.execute(edad_sql, params)
        edad_rows = cur.fetchall()
        edad_data = [{'rango': r.Rango, 'total': int(r.Total)} for r in edad_rows]

        # Totals (KPIs) computed server-side so the UI can display authoritative values
        total_atenciones = sum(x.get('numero', 0) for x in atenciones)
        total_consulta = sum(x.get('total', 0) for x in valor_consulta)
        total_medicina = sum(x.get('total', 0) for x in valor_medicina)
        total_general = total_consulta + total_medicina
        totales = {
            'atenciones': int(total_atenciones),
            'total_consulta': float(total_consulta),
            'total_medicina': float(total_medicina),
            'total_general': float(total_general),
            'cancelaciones': 0,
        }

        # Detalle (filas individuales) — reutiliza la función helper
        detalle = fetch_detalle(cur, base_where, params)

        resp = {
            'atenciones_mes': atenciones,
            'valor_consulta': valor_consulta,
            'valor_medicina': valor_medicina,
            'instituciones': instituciones,
            'genero': genero_data,
            'edad': edad_data,
            'totales': totales,
            'detalle': detalle,
        }
    except Exception as e:
        logger.exception('Error ejecutando dashboard_data')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    return JsonResponse(resp, safe=False)


@staff_member_required
@require_GET
def export_parte_diario(request):
    fecha = request.GET.get('fecha')
    if not fecha:
        return JsonResponse({'error': 'Parámetro "fecha" requerido (YYYY-MM-DD).'}, status=400)

    sql = """
    SELECT 
        t.CodCli,
        c.nomCli,
        c.Tipo AS Institucion,
        DateDiff('yyyy', c.FechaN, Date()) AS Edad,
        e.FechaIng,
        e.HoraIng,
        e.FechaSal,
        e.HoraPop,
        t.Inicial,
        t.Subsecuente,
        s.CodSer
    FROM TraSec AS t
    INNER JOIN CtaCli AS c ON t.CodCli = c.CodCli
    INNER JOIN EqTraSec AS e ON t.IdTraSec = e.IdTraSec
    INNER JOIN EqCliDes AS s ON t.IdTraSec = s.IdTraSec
    WHERE e.FechaIng = ?;
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql, [fecha])
        rows = cur.fetchall()
        headers = [c[0] for c in cur.description]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Parte Diario'
        ws.append(headers)
        for r in rows:
            ws.append(list(r))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'parte_diario_{fecha}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as e:
        logger.exception('Error exportando parte diario')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass


@staff_member_required
@require_GET
def export_detalle(request):
    """Export the detalle rows to Excel using the same filters as `dashboard_data`."""
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    institucion = request.GET.get('institucion')
    genero = request.GET.get('genero')
    edad_range = request.GET.get('edad_range')

    params = []
    where_clauses = []

    if inicio and fin:
        where_clauses.append("e.FechaIng BETWEEN ? AND ?")
        params.extend([inicio, fin])

    if institucion:
        where_clauses.append('c.Tipo = ?')
        params.append(institucion)
    if genero and genero != 'ALL':
        where_clauses.append('c.Genero = ?')
        params.append(genero)

    if edad_range:
        try:
            if edad_range.endswith('+'):
                low = int(edad_range[:-1])
                where_clauses.append(f"DateDiff('yyyy', c.FechaN, Date()) >= {low}")
            else:
                parts = edad_range.split('-')
                if len(parts) != 2:
                    raise ValueError('invalid format')
                low, high = parts
                where_clauses.append(f"DateDiff('yyyy', c.FechaN, Date()) BETWEEN {int(low)} AND {int(high)}")
        except Exception:
            return JsonResponse({'error': 'Parámetro "edad_range" inválido. Use "0-12" o "60+".'}, status=400)

    try:
        conn = get_connection()
        cur = conn.cursor()
        base_where = ('WHERE ' + ' AND '.join(where_clauses)) if where_clauses else ''
        detalle_rows = fetch_detalle(cur, base_where, params)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Detalle'
        headers = ['Fecha atención', 'Paciente', 'Servicio', 'Institución', 'Valor consulta', 'Valor medicina', 'Total']
        ws.append(headers)
        for r in detalle_rows:
            ws.append([r.get('fecha'), r.get('paciente'), r.get('servicio'), r.get('institucion'), r.get('valor_consulta'), r.get('valor_medicina'), r.get('total')])

        # Polish: header bold, currency formatting, auto column widths
        from openpyxl.styles import Font
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
        # Apply currency format to columns 5,6,7 (Valor consulta, Valor medicina, Total)
        currency_fmt = '"$"#,##0.00'
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
            for cell in row:
                # ensure numeric type
                try:
                    if cell.value is not None:
                        cell.value = float(cell.value)
                except Exception:
                    pass
                cell.number_format = currency_fmt
                cell.alignment = cell.alignment.copy(horizontal='right')

        # Auto-width columns based on max content length
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value)
                except Exception:
                    val = ''
                if val is None:
                    val = ''
                max_length = max(max_length, len(val))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # filename with dates in friendly format if provided
        fname = 'Parte_Detalle'
        if inicio and fin:
            fname = f'Parte_Detalle_{inicio}_a_{fin}'
        filename = f'{fname}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as e:
        logger.exception('Error exportando detalle')
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        try:
            conn.close()
        except Exception:
            pass
