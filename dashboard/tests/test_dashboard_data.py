import json
import unittest
from types import SimpleNamespace
from unittest.mock import patch
from django.test import RequestFactory
from django.http import JsonResponse

class MockRow:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


class MockCursor:
    def __init__(self):
        self.executed = []
        self.last_sql = None
        self.last_params = None
        self.description = []

    def execute(self, sql, params=None):
        self.executed.append((sql, params or []))
        self.last_sql = sql
        self.last_params = params or []
        if "DatePart('m', e.FechaIng) AS Mes" in sql:
            self.description = [('Mes',), ('NumeroAtenciones',)]
        elif 'SUM(c.ValorConsulta)' in sql:
            self.description = [('Mes',), ('TotalConsulta',)]
        elif 'SUM(m.TotalMedicina)' in sql:
            self.description = [('Mes',), ('TotalMedicina',)]
        elif 'c.Tipo AS Institucion' in sql:
            self.description = [('Institucion',), ('TotalAtenciones',)]
        elif 'c.Genero' in sql and 'COUNT' in sql:
            self.description = [('Genero',), ('Total',)]
        elif "IIF(DateDiff('yyyy'" in sql:
            self.description = [('Rango',), ('Total',)]
        elif "e.FechaIng AS Fecha" in sql or "m.CodSer" in sql:
            self.description = [('Fecha',), ('Paciente',), ('Servicio',), ('Institucion',), ('ValorConsulta',), ('ValorMedicina',)]

    def fetchall(self):
        sql = self.last_sql or ''
        if "e.FechaIng AS Fecha" in sql or "m.CodSer" in sql:
            return [MockRow(Fecha='2025-01-15', Paciente='Juan Perez', Servicio='001', Institucion='Publica', ValorConsulta=50.0, ValorMedicina=10.0), MockRow(Fecha='2025-01-16', Paciente='Ana Ruiz', Servicio='002', Institucion='Privada', ValorConsulta=100.0, ValorMedicina=0.0)]
        if "DatePart('m', e.FechaIng) AS Mes" in sql and 'NumeroAtenciones' in sql:
            return [MockRow(Mes=1, NumeroAtenciones=5), MockRow(Mes=2, NumeroAtenciones=7)]
        if 'SUM(c.ValorConsulta)' in sql:
            return [MockRow(Mes=1, TotalConsulta=100.0), MockRow(Mes=2, TotalConsulta=150.0)]
        if 'SUM(m.TotalMedicina)' in sql:
            return [MockRow(Mes=1, TotalMedicina=20.0), MockRow(Mes=2, TotalMedicina=30.0)]
        if 'c.Tipo AS Institucion' in sql:
            return [MockRow(Institucion='Publica', TotalAtenciones=8), MockRow(Institucion='Privada', TotalAtenciones=4)]
        if 'c.Genero' in sql and 'COUNT' in sql:
            return [MockRow(Genero='M', Total=7), MockRow(Genero='F', Total=5)]
        if "IIF(DateDiff('yyyy'" in sql:
            return [MockRow(Rango='18-30', Total=6), MockRow(Rango='31-59', Total=6)]
        return []


class MockConn:
    def __init__(self):
        self.cursor_obj = MockCursor()

    def cursor(self):
        return self.cursor_obj

    def close(self):
        pass


class DashboardDataUnitTests(unittest.TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    @patch('dashboard.views.get_connection')
    def test_filters_applied_dates_gender_institucion_age(self, mock_get_conn):
        """Ensure consolidated view composes queries including date/gender/institucion/age filters."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/', {
            'inicio': '2025-01-01', 'fin': '2025-12-31',
            'genero': 'M', 'institucion': 'Publica', 'edad_range': '18-30'
        })
        # bypass decorators to call original implementation
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        # attach a minimal user object (not hitting DB)
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        resp = orig(req)
        self.assertIsInstance(resp, JsonResponse)
        data = json.loads(resp.content)
        self.assertIn('atenciones_mes', data)
        cur = mock_get_conn.return_value.cursor_obj
        executed_sqls = [s for s, p in cur.executed]
        joined = '\n'.join(executed_sqls)
        self.assertIn("e.FechaIng BETWEEN ? AND ?", joined)
        self.assertIn("c.Genero = ?", joined)
        self.assertIn("c.Tipo = ?", joined)
        self.assertIn("DateDiff('yyyy', c.FechaN, Date()) BETWEEN 18 AND 30", joined)

    @patch('dashboard.views.get_connection')
    def test_totales_in_response(self, mock_get_conn):
        """Ensure the consolidated endpoint returns server-side KPIs under 'totales'."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig(req)
        self.assertIsInstance(resp, JsonResponse)
        data = json.loads(resp.content)
        self.assertIn('totales', data)
        tot = data['totales']
        # Based on MockCursor returns: atenciones 5+7=12, consulta 100+150=250, medicina 20+30=50
        self.assertEqual(tot.get('atenciones'), 12)
        self.assertEqual(tot.get('total_consulta'), 250.0)
        self.assertEqual(tot.get('total_medicina'), 50.0)
        self.assertEqual(tot.get('total_general'), 300.0)

    @patch('dashboard.views.get_connection')
    def test_export_detalle_generates_xlsx(self, mock_get_conn):
        """Export endpoint should return a valid XLSX reflecting the detalle rows."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/export/detalle/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.export_detalle.__wrapped__.__wrapped__
        resp = orig(req)
        # HttpResponse with xlsx
        self.assertEqual(resp.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resp['Content-Type'])
        # open the workbook from response content
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        # header names in Spanish, same order as UI table
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, ['Fecha atención', 'Paciente', 'Servicio', 'Institución', 'Valor consulta', 'Valor medicina', 'Total'])
        # check one data row
        self.assertEqual(ws.max_row, 3)  # header + 2 mock rows
        first_data = [ws.cell(row=2, column=c).value for c in range(1,8)]
        self.assertEqual(first_data[1], 'Juan Perez')
        self.assertEqual(first_data[4], 50.0)
        # header is bold
        for h in ws[1]:
            self.assertTrue(h.font.bold)
        # numeric columns have currency format and are numeric
        self.assertIn('0.00', ws.cell(row=2, column=5).number_format)
        self.assertIsInstance(ws.cell(row=2, column=5).value, (int, float))
        self.assertIn('0.00', ws.cell(row=2, column=7).number_format)
        # filename contains the chosen pattern
        self.assertIn('Parte_Detalle_2025-01-01_a_2025-12-31.xlsx', resp['Content-Disposition'])

    @patch('dashboard.views.get_connection')
    def test_export_detalle_zero_rows(self, mock_get_conn):
        """Export endpoint should return a valid XLSX even when there are no rows (only header)."""
        class ConnZero:
            def __init__(self):
                self._c = MockCursor()
            def cursor(self):
                cur = self._c
                def fetchall_empty():
                    return []
                cur.fetchall = fetchall_empty
                return cur
            def close(self):
                pass

        mock_get_conn.return_value = ConnZero()
        req = self.factory.get('/dashboard/export/detalle/')
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.export_detalle.__wrapped__.__wrapped__
        resp = orig(req)
        self.assertEqual(resp.status_code, 200)
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        # header present and no data rows
        self.assertEqual(ws.max_row, 1)

    @patch('dashboard.views.get_connection')
    def test_detalle_in_response(self, mock_get_conn):
        """Consolidated endpoint should include a 'detalle' array with rows for the period."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig(req)
        self.assertIsInstance(resp, JsonResponse)
        data = json.loads(resp.content)
        self.assertIn('detalle', data)
        detalle = data['detalle']
        self.assertIsInstance(detalle, list)
        # check first row keys and values
        first = detalle[0]
        self.assertEqual(first.get('paciente'), 'Juan Perez')
        self.assertEqual(first.get('servicio'), '001')
        self.assertEqual(first.get('valor_consulta'), 50.0)
        self.assertEqual(first.get('valor_medicina'), 10.0)
        self.assertEqual(first.get('total'), 60.0)

    @patch('dashboard.views.get_connection')
    def test_cache_hit_miss(self, mock_get_conn):
        """Validate cache wrapper reduces calls to the DB for identical requests."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        # get the function wrapped by cache_page (two unwraps gives cache wrapper)
        cache_wrapped = dviews.dashboard_data.__wrapped__.__wrapped__
        # first call -> should invoke DB
        resp1 = cache_wrapped(req)
        self.assertIsInstance(resp1, JsonResponse)
        self.assertTrue(mock_get_conn.called)
        mock_get_conn.reset_mock()
        # second identical call -> should be served from cache, no DB call
        resp2 = cache_wrapped(req)
        self.assertIsInstance(resp2, JsonResponse)
        self.assertFalse(mock_get_conn.called)

    @patch('dashboard.views.get_connection')
    def test_empty_response_structures(self, mock_get_conn):
        """When DB returns no rows, the endpoint still returns the expected array keys (possibly empty)."""
        class ConnEmpty:
            def __init__(self):
                self._c = MockCursor()
            def cursor(self):
                cur = self._c
                def fetchall_empty():
                    return []
                cur.fetchall = fetchall_empty
                cur.description = []
                return cur
            def close(self):
                pass

        mock_get_conn.return_value = ConnEmpty()
        req = self.factory.get('/dashboard/api/data/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig(req)
        self.assertIsInstance(resp, JsonResponse)
        data = json.loads(resp.content)
        # Keys should be present and be lists (empty)
        for k in ['atenciones_mes','valor_consulta','valor_medicina','instituciones','genero','edad','detalle']:
            self.assertIn(k, data)
            self.assertIsInstance(data[k], list)

        # different querystring -> cache miss -> DB called
        req2 = self.factory.get('/dashboard/api/data/', {'inicio': '2025-02-01', 'fin': '2025-12-31'})
        req2.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        cache_wrapped = dviews.dashboard_data.__wrapped__.__wrapped__
        resp3 = cache_wrapped(req2)
        self.assertIsInstance(resp3, JsonResponse)
        self.assertTrue(mock_get_conn.called)


if __name__ == '__main__':
    unittest.main()


class DashboardDataEdgeCases(unittest.TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    @patch('dashboard.views.get_connection')
    def test_zero_results_returns_empty_lists(self, mock_get_conn):
        """When queries return no rows, endpoint returns empty lists (no crash)."""
        class ConnZero:
            def __init__(self):
                self._c = MockCursor()
            def cursor(self):
                # force fetchall to return [] regardless of SQL
                cur = self._c
                def fetchall_empty():
                    return []
                cur.fetchall = fetchall_empty
                return cur
            def close(self):
                pass

        mock_get_conn.return_value = ConnZero()
        req = self.factory.get('/dashboard/api/data/')
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig(req)
        self.assertIsInstance(resp, JsonResponse)
        data = json.loads(resp.content)
        # expect keys present and empty lists
        for k in ('atenciones_mes', 'valor_consulta', 'valor_medicina', 'instituciones', 'genero', 'edad'):
            self.assertIn(k, data)
            self.assertTrue(isinstance(data[k], list))
        # and totales should be present with zeros
        self.assertIn('totales', data)
        tot = data['totales']
        self.assertEqual(tot.get('atenciones'), 0)
        self.assertEqual(tot.get('total_general'), 0.0)
        # detalle should be present but empty
        self.assertIn('detalle', data)
        self.assertIsInstance(data['detalle'], list)
        self.assertEqual(len(data['detalle']), 0)

    @patch('dashboard.views.get_connection')
    def test_invalid_age_range_returns_error(self, mock_get_conn):
        """Malformed age_range should not crash server; expect error response (handled or 500)."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/', {'edad_range': 'bad-format'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig(req)
        # Accept either successful handling or an error - ensure it's a HTTP response
        self.assertTrue(hasattr(resp, 'status_code'))
        self.assertIn(resp.status_code, (200, 400, 500))

    @patch('dashboard.views.get_connection')
    def test_missing_filters_returns_data(self, mock_get_conn):
        """No filters provided: endpoint should still return a response with datasets."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/')
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        from dashboard import views as dviews
        orig = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig(req)
        self.assertIsInstance(resp, JsonResponse)
        data = json.loads(resp.content)
        self.assertIn('atenciones_mes', data)
