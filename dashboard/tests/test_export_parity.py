import json
from types import SimpleNamespace
from unittest.mock import patch
from django.test import RequestFactory, TestCase
from openpyxl import load_workbook
from io import BytesIO

from dashboard import views as dviews
from dashboard.tests.test_dashboard_data import MockConn, MockCursor


class DashboardExportParityTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    @patch('dashboard.views.get_connection')
    def test_api_and_excel_parity(self, mock_get_conn):
        """The API 'detalle' must match the rows inside the exported XLSX exactly."""
        mock_get_conn.return_value = MockConn()
        req = self.factory.get('/dashboard/api/data/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        orig_data = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig_data(req)
        data = json.loads(resp.content)
        detalle = data.get('detalle', [])

        # export
        req2 = self.factory.get('/dashboard/export/detalle/', {'inicio': '2025-01-01', 'fin': '2025-12-31'})
        req2.user = SimpleNamespace(is_staff=True, is_active=True)
        orig_export = dviews.export_detalle.__wrapped__.__wrapped__
        xresp = orig_export(req2)
        self.assertEqual(xresp.status_code, 200)
        wb = load_workbook(filename=BytesIO(xresp.content))
        ws = wb.active
        excel_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            excel_rows.append({
                'fecha': row[0],
                'paciente': row[1],
                'servicio': row[2],
                'institucion': row[3],
                'valor_consulta': float(row[4] or 0),
                'valor_medicina': float(row[5] or 0),
                'total': float(row[6] or 0),
            })

        self.assertEqual(len(detalle), len(excel_rows))
        self.assertEqual(detalle, excel_rows)

    @patch('dashboard.views.get_connection')
    def test_zero_rows_parity(self, mock_get_conn):
        """When there are no detalle rows, API should return empty list and XLSX should contain only header."""
        class ConnZero:
            def cursor(self):
                cur = MockCursor()
                def fetchall_empty():
                    return []
                cur.fetchall = fetchall_empty
                return cur
            def close(self):
                pass

        mock_get_conn.return_value = ConnZero()
        req = self.factory.get('/dashboard/api/data/')
        req.user = SimpleNamespace(is_staff=True, is_active=True)
        orig_data = dviews.dashboard_data.__wrapped__.__wrapped__.__wrapped__
        resp = orig_data(req)
        data = json.loads(resp.content)
        self.assertIn('detalle', data)
        self.assertEqual(len(data['detalle']), 0)

        req2 = self.factory.get('/dashboard/export/detalle/')
        req2.user = SimpleNamespace(is_staff=True, is_active=True)
        orig_export = dviews.export_detalle.__wrapped__.__wrapped__
        xresp = orig_export(req2)
        self.assertEqual(xresp.status_code, 200)
        wb = load_workbook(filename=BytesIO(xresp.content))
        ws = wb.active
        # only header row
        self.assertEqual(ws.max_row, 1)
